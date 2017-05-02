# -*- coding: utf-8 -*-
"""
Created on Mon May  1 00:28:28 2017

@author: Jon-Michael
"""

from openpyxl import load_workbook
from random import randint
from copy import copy, deepcopy
import numpy as np
import requests
import json
import time


workbook = load_workbook("seedStatsData.xlsx", data_only=True)
worksheet = workbook.worksheets[0]

row_count = worksheet.max_row


nte = True
row = 0

data = []

for row in worksheet.iter_rows():
    holder = []
    for cell in row:
        holder.append(cell.internal_value)

    data.append(holder)

sIds = []

for row in range(0, len(data)):
    sIds.append(data[row][0])

testMatchDB = []

for i in range (0, 499):
    time.sleep(2)
    sumID = sIds[i * 2]    
    matchlistUrl = "https://na1.api.riotgames.com/lol/match/v3/matchlists/by-account/"+ str(sumID) + "?api_key=836619ee-c877-45d9-b718-ab0eea4ed172"
    response = requests.get(matchlistUrl)
    matchlist = response.json()
    try:
        totGames = matchlist['totalGames']
    except:
        totGames = 0
    
    if len(testMatchDB) > 2000:
        break
    
    if totGames > 0:
        for stuff in range(0, totGames):
            testMatchDB.append(matchlist['matches'][stuff]['gameId'])
            
    
    
    