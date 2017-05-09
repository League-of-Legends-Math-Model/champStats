# -*- coding: utf-8 -*-
"""
Created on Sun Apr 23 10:04:24 2017

@author: Jon-Michael
"""

from openpyxl import load_workbook
from random import randint
from copy import copy, deepcopy
import numpy as np
import math
import matplotlib.pyplot as plt
import pylab


workbook = load_workbook("seeddata.xlsx", data_only=True)
worksheet = workbook.worksheets[0]

rowCount = worksheet.max_row

data = []

for row in worksheet.iter_rows():
    holder = []
    for cell in row:
        holder.append(cell.internal_value)

    data.append(holder)
    
sortedData = [[],[],[],[],[],[],[]]

rows = len(data)
cols = len(data[0])

for i in range(0, rows):
    if int(data[i][3]) > 0:
        sortedData[int(data[i][3]) - 1].append(data[i])
        
#Scatters

#Rank vs CS/min
ranks = []
csminRank = []
position = []
gametime = []
csTotal = []
damageEff = []
durab = []
wardsPlaced = []
wardsKilled = []
killPart = []

for i in range(0, len(sortedData)):
    for j in range(0, len(sortedData[i])):
        ranks.append(i)
        csminRank.append(sortedData[i][j][36])
        damageEff.append(sortedData[i][j][47])
        durab.append(sortedData[i][j][48])
        gametime.append(sortedData[i][j][1])
        wardsPlaced.append(sortedData[i][j][25])
        wardsKilled.append(sortedData[i][j][26])
        killPart.append(sortedData[i][j][43])
        csTotal.append(sortedData[i][j][20] + sortedData[i][j][21])
        lane = int(sortedData[i][j][4])
        role = int(sortedData[i][j][5])
        if lane == 3:
            if role < 4:
                lane = 4
        position.append(lane)
        
plt.scatter(csminRank, ranks)

plt.show()

# Remove irrelevant data

cumulative = [ranks, position, gametime, csTotal, damageEff, durab, wardsPlaced, wardsKilled, killPart]

better = []

for i in range(len(ranks)):
    if gametime[i] > 500:
        if position[i] < 4:
            if csTotal[i] > 25:
                helper = []
                for j in range(0, len(cumulative)):
                    helper.append(cumulative[j][i])
                better.append(helper)
        else:
            helper = []
            for j in range(0, len(cumulative)):
                helper.append(cumulative[j][i])
            better.append(helper)
    

"""
Ranks and Colors
Bronze: red
Silver: green
Gold: blue
Platinum: purple
Diamond: magenta
Master: black
Challenger: gold
"""
#Rank Pos vs. CS/min

filtered = [[],[],[],[],[],[],[],[],[],[],[],[],[],[],[]]

partitioned = [[],[],[],[],[],[],[]]

for i in range(0, len(partitioned)):
    for j in range(0, 10):
        partitioned[i].append([])

colors = ['red', 'green', 'blue', 'purple', 'magenta', 'black', 'gold']


for i in range(0, len(better)):
    sortTo = better[i][1] * 3
    psortTo = better[i][1] * 2
    filtered[sortTo].append(better[i][2])
    filtered[sortTo+1].append(better[i][3])
    filtered[sortTo+2].append(colors[better[i][0]])
    partitioned[better[i][0]][psortTo].append(better[i][2])
    partitioned[better[i][0]][psortTo + 1].append(better[i][3])

    
for i in range(0, 5):
    plt.scatter(filtered[i*3], filtered[i*3 + 1], color = filtered[i*3 + 2])
    plt.show()

ranker = ["Bronze", "Silver", "Gold", "Platinum", "Diamond", "Master", "Challenger"]
posits = ["Top", "Jungle", "Mid", "ADC", "Support"]

for i in range(0, len(partitioned)):
    for j in range(0, 5):
        print("Rank: "+ranker[i] + " Position: " + posits[j])
        if len(partitioned[i][j*2]) > 0:
            plt.scatter(partitioned[i][j*2],partitioned[i][j*2 + 1],color = colors[i])
            plt.show()

# REGRESSION ON THE INDIVIDUAL ROLES/RANKS

csDatapoints = [[],[],[],[],[],[],[]]
csResults = [[],[],[],[],[],[],[]]
csTestdatapoints = [[],[],[],[],[],[],[]]
csTestresults = [[],[],[],[],[],[],[]]

for i in range(0, len(csDatapoints)):
    csData = [[],[],[],[],[]]
    csRes = [[],[],[],[],[]]
    csTesD = [[],[],[],[],[]]
    csTesR = [[],[],[],[],[]]
    for j in range(0, 5):
        if len(partitioned[i][j*2]) > 0:
           for k in range(0, len(partitioned[i][j*2])):
               if k < len(partitioned[i][j*2]) /2:
                   csData[j].append(partitioned[i][j*2 + 1][k])
                   csRes[j].append(partitioned[i][j*2][k])
               else:
                   csTesD[j].append(partitioned[i][j*2 + 1][k])
                   csTesR[j].append(partitioned[i][j*2][k])
            
    csDatapoints[i].append(csData)
    csResults[i].append(csRes)
    csTestdatapoints[i].append(csTesD)
    csTestresults[i].append(csTesR)

csRegForm = [[],[],[],[],[],[],[]]
for i in range(0, 5):
    for j in range(0, 5):
        if len(csDatapoints[i][0][j]) > 0:
            info = np.zeros(shape = (len(csDatapoints[i][0][j]), 2))
            for k in range(0, len(csDatapoints[i][0][j])):
                info[k][0] = 1
                info[k][1] = csDatapoints[i][0][j][k]
            ytp = deepcopy(info)
            ytpTrans = deepcopy(ytp.transpose())
            mult = ytpTrans.dot(ytp)
            invs = np.linalg.inv(mult)
            ytpPlus = invs.dot(ytpTrans)
            coeff = ytpPlus.dot(csResults[i][0][j])
            betCoeff = coeff.tolist()
            csRegForm[i].append(betCoeff)

guesses = [[],[],[],[],[],[],[]]

for i in range(0, 5):
    for j in range(0, 5):
        for k in range(0, len(csTestdatapoints[i][0][j])):
            x = csTestdatapoints[i][0][j][k]
            y = csTestresults[i][0][j][k]
            distances = []
            for p in range(0, 5):
                slope = csRegForm[p][j][1]
                yinter = csRegForm[p][j][0]
                nslope = -1/slope
                nyinter = y - nslope * x
                interx = (nyinter - yinter) / (slope - nslope)
                distance = math.sqrt(math.pow((interx - x),2) + math.pow((interx * slope + yinter - y),2))
                distances.append(distance)
            shortest = distances[0]
            choice = 0
            for p in range(0, len(distances)):
                if distances[p] < shortest:
                    shortest = distances[p]
                    choice = p
            guesses[i].append(choice + 1)

#ACCURACY

results = []

for i in range(0, 5):
    correct = 0
    for j in range(0, len(guesses[i])):
        if guesses[i][j] == i:
            correct += 1
    results.append(correct/len(guesses[i]))

    
# Damage Effectivenss vs. Durability

ddDatapoints = [[],[],[],[],[],[],[]]

for i in range(0, len(ddDatapoints)):
    for j in range(0, 10):
        ddDatapoints[i].append([])

for i in range(0, len(better)):
    sortTo = better[i][1] * 2
    rankedF = better[i][0]
    ddDatapoints[rankedF][sortTo].append(better[i][4])
    ddDatapoints[rankedF][sortTo + 1].append(better[i][5])

# SCATTERS

for i in range(0, len(ddDatapoints)):
    for j in range(0, 5):
        print("Rank: " + ranker[i] + " Position: " + posits[j])
        plt.scatter(ddDatapoints[i][j*2], ddDatapoints[i][j*2 + 1], color = colors[i])
        plt.show()

#COMBINED SCATTERS

for j in range(0, 5):
    print("Position: " + posits[j])
    for i in range(0, len(ddDatapoints)):
        plt.scatter(ddDatapoints[i][j*2], ddDatapoints[i][j*2 + 1], color = colors[i])
    plt.show()
    
# Wards Placed and Wards Killed

wpDatapoints = [[],[],[],[],[],[],[]]
wkDatapoints = [[],[],[],[],[],[],[]]

for i in range(0, len(wpDatapoints)):
    for j in range(0, 10):
        wpDatapoints[i].append([])
        wkDatapoints[i].append([])

for i in range(0, len(better)):
    sortedTo = better[i][1] * 2
    rankedF = better[i][0]
    #gametime is 2
    wpDatapoints[rankedF][sortedTo].append(better[i][2])
    wpDatapoints[rankedF][sortedTo + 1].append(better[i][6])
    wkDatapoints[rankedF][sortedTo].append(better[i][2])
    wkDatapoints[rankedF][sortedTo + 1].append(better[i][7])

# SCATTERS

for j in range(0, 5):
    print("Position: " + posits[j])
    for i in range(0, len(wpDatapoints)):
        plt.scatter(wpDatapoints[i][j*2], wpDatapoints[i][j*2 + 1], color = colors[i])
    plt.show()

for j in range(0, 5):
    print("Position: " + posits[j])
    for i in range(0, len(wkDatapoints)):
        plt.scatter(wkDatapoints[i][j*2], wkDatapoints[i][j*2 + 1], color = colors[i])
    plt.show()

#Kill Participation

kpDatapoints = [[],[],[],[],[],[],[]]

for i in range(0, len(kpDatapoints)):
    for j in range(0, 10):
        kpDatapoints[i].append([])
for i in range(0, len(better)):
    sortedTo = better[i][1] * 2
    rankedF = better[i][0]
    kpDatapoints[rankedF][sortedTo].append(better[i][8])
    kpDatapoints[rankedF][sortedTo + 1].append(better[i][0])

# SCATTERS

for j in range(0, 5):
    print("Position: " + posits[j])
    for i in range(0, len(kpDatapoints)):
        plt.scatter(kpDatapoints[i][j*2], kpDatapoints[i][j*2 + 1], color = colors[i])
    plt.show()

"""
halves = [0, 0, 0, 0, 0, 0, 0]

for i in range(0, len(sortedData)):
    halves[i] = round(len(sortedData[i])/2)

trainData = [[],[],[],[],[],[],[]]
testData = [[],[],[],[],[],[],[]]

for i in range(0, len(sortedData)):
    for j in range(0, len(sortedData[i])):
        if j < halves[i]:
            testData[i].append(sortedData[i][j])
        else:
            trainData[i].append(sortedData[i][j])

# LINEAR REGRESSION ON ALL NEW DATA POINTS

ranks = [[],[],[],[],[]]
datapoints = [[],[],[],[],[]]
oDatapoints = [[],[],[],[],[]]
for i in range(0, len(trainData)):
    for j in range(0, len(trainData[i])):
        lane = int(trainData[i][j][4])
        role = int(trainData[i][j][5])
        if lane == 3:
            if role < 4:
                lane = 4
        ranks[lane].append(i)
        holder = []
        holder.append(trainData[i][j][9])
        holder.append(trainData[i][j][27])
        for k in range(36, 44):
            holder.append(trainData[i][j][k])
        for k in range(45, 54):
            holder.append(trainData[i][j][k])
        sansRoles = deepcopy(holder)
        datapoints[lane].append(sansRoles)
        for k in range(54, 57):
            holder.append(trainData[i][j][k])
        oDatapoints[lane].append(holder)

reg = []

# REGRESSION ON DATA WITHOUT ROLE PREDICTIONS

for i in range(0, len(ranks)):
    if len(datapoints[i]) > 0:
        info = np.ndarray(shape = (len(datapoints[i]), len(datapoints[i][0]) + 1))
        for rd in range(0, len(datapoints[i])):
            info[rd][0] = 1
            for dr in range(0, len(datapoints[i][0])):
                info[rd][dr + 1] = datapoints[i][rd][dr]
        ytp = deepcopy(info)
        ytpTrans = deepcopy(ytp.transpose())
        mult = ytpTrans.dot(ytp)
        invs = np.linalg.inv(mult)
        ytpPlus = invs.dot(ytpTrans)
        coeff = ytpPlus.dot(ranks[i])
        betCoeff = coeff.tolist()
        reg.append(betCoeff)

# REGRESSION ON DATA WITH ROLE PREDICTIONS
"""
"""
oreg = []

for i in range(0, len(ranks)):
    if len(oDatapoints[i]) > 0:
        info = np.ndarray(shape = (len(oDatapoints[i]), len(oDatapoints[i][0]) + 1))
        for rd in range(0, len(oDatapoints[i])):
            info[rd][0] = 1
            for dr in range(0, len(oDatapoints[i][0])):
                info[rd][dr + 1] = oDatapoints[i][rd][dr]
        ytp = deepcopy(info)
        ytpTrans = deepcopy(ytp.transpose())
        mult = ytpTrans.dot(ytp)
        invs = np.linalg.inv(mult)
        ytpPlus = invs.dot(ytpTrans)
        coeff = ytpPlus.dot(ranks[i])
        betCoeff = coeff.tolist()
        oreg.append(betCoeff)
"""
"""
# PREP TEST DATA

testRanks = [[],[],[],[],[]]
testpoints = [[],[],[],[],[]]
#oTestpoints = [[],[],[],[],[]]
for i in range(0, len(testData)):
    for j in range(0, len(testData[i])):
        lane = int(testData[i][j][4])
        role = int(testData[i][j][5])
        if lane == 3:
            if role < 4:
                lane = 4
        testRanks[lane].append(i)
        holder = []
        holder.append(testData[i][j][9])
        holder.append(testData[i][j][27])
        for k in range(36, 44):
            holder.append(testData[i][j][k])
        for k in range(45, 54):
            holder.append(testData[i][j][k])
        sansRoles = deepcopy(holder)
        testpoints[lane].append(sansRoles)
        for k in range(54, 57):
            holder.append(testData[i][j][k])
        #oTestpoints[lane].append(holder)

# APPLY LINEAR REGRESSION - REG

roundPred = [[],[],[],[],[]]
floorPred = [[],[],[],[],[]]
ceilPred = [[],[],[],[],[]]
for alpha in range(0, len(testpoints)):
    if len(testpoints[alpha]) > 0:
        for j in range(0, len(testpoints[alpha])):
            expected = reg[alpha][0]
            for k in range(0, len(testpoints[alpha][j])):
                expected += testpoints[alpha][j][k] * reg[alpha][k + 1]
            about = round(expected)
            down = math.floor(expected)
            up = math.ceil(expected)
            roundPred[alpha].append(about)
            floorPred[alpha].append(down)
            ceilPred[alpha].append(up)

# APPLY LINEAR PROGRESSION - OREG
"""
"""
oroundPred = [[],[],[],[],[]]
ofloorPred = [[],[],[],[],[]]
oceilPred = [[],[],[],[],[]]
for i in range(0, len(oTestpoints)):
    if len(oTestpoints[i]) > 0:
        for j in range(0, len(oTestpoints[i])):
            expected = oreg[i][0]
            for k in range(0, len(oTestpoints[i][j])):
                expected += oTestpoints[i][j][k] * oreg[i][k + 1]
            about = round(expected)
            down = math.floor(expected)
            up = math.ceil(expected)
            oroundPred[i].append(about)
            ofloorPred[i].append(down)
            oceilPred[i].append(up)
"""
"""
# ACCURACY

regRound = []
regFloor = []
regCeil = []
#oregRound = []
#oregFloor = []
#oregCeil = []

for i in range(0, len(roundPred)):
    rracc = 0
    rfacc = 0
    rcacc = 0
    oracc = 0
    ofacc = 0
    ocacc = 0
    cases = len(roundPred[i])
    if cases > 0:
        for j in range(0, cases):
            if roundPred[i][j] == testRanks[i][j]:
                rracc += 1
            #if oroundPred[i][j] == testRanks[i][j]:
            #    oracc += 1
            if floorPred[i][j] == testRanks[i][j]:
                rfacc += 1
            #if ofloorPred[i][j] == testRanks[i][j]:
            #    ofacc += 1
            if ceilPred[i][j] == testRanks[i][j]:
                rcacc += 1
            #if oceilPred[i][j] == testRanks[i][j]:
            #    ocacc += 1
        rracc = rracc / cases
        #oracc = oracc / cases
        rfacc = rfacc / cases
        #ofacc = ofacc / cases
        rcacc = rcacc / cases
        #ocacc = ocacc / cases
    regRound.append(rracc)
    regFloor.append(rfacc)
    regCeil.append(rcacc)
    #oregRound.append(oracc)
    #oregFloor.append(ofacc)
    #oregCeil.append(ocacc)


"""