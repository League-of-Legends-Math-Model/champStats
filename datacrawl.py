# -*- coding: utf-8 -*-
"""
Created on Fri May  5 22:17:54 2017

@author: Jon-Michael
"""

from openpyxl import load_workbook
from copy import copy, deepcopy
import numpy as np
import requests
import json
import time

workbook = load_workbook("sIdsRanks.xlsx", data_only=True)
worksheet = workbook.worksheets[0]

rowCount = worksheet.max_row

br = []
sv = []
gd = []
pl = []
di = []
ma = []
ch = []

for row in worksheet.iter_rows():
    holder = []
    for cell in row:
        holder.append(cell.internal_value)
    
    if holder[1] == 1:
        br.append(holder[0])
    
    if holder[1] == 2:
        sv.append(holder[0])
    
    if holder[1] == 3:
        gd.append(holder[0])
        
    if holder[1] == 4:
        pl.append(holder[0])
    
    if holder[1] == 5:
        di.append(holder[0])
    
    if holder[1] == 6:
        ma.append(holder[0])
    
    if holder[1] == 7:
        ch.append(holder[0])

bPrior = len(br)/10000
sPrior = len(sv)/10000
gPrior = len(gd)/10000
pPrior = len(pl)/10000
dPrior = len(di)/10000
mPrior = len(ma)/10000
cPrior = len(ch)/10000

crossSection = [len(br), len(sv), len(gd), len(pl), len(di), len(ma), len(ch)] 

rankedPop = 0
for i in range(0, len(crossSection)):
    rankedPop += crossSection[i]

for i in range(0, len(crossSection)):
    crossSection[i] = crossSection[i] / rankedPop

caps = [0, 0, 0, 0, 0, 0, 0]

for i in range(0, len(caps)):
    caps[i] = round(crossSection[i] * 2000)

sumIds = [br, sv, gd, pl, di, ma, ch]

chGames = []
sumName = "Pobelter"
chstuff = "https://na1.api.riotgames.com/lol/summoner/v3/summoners/by-name/"+sumName+"?api_key=836619ee-c877-45d9-b718-ab0eea4ed172"
response = requests.get(chstuff)
sumstuff = response.json()
sumId = sumstuff['accountId']
matchUrl = "https://na1.api.riotgames.com/lol/match/v3/matchlists/by-account/"+str(sumId)+"?api_key=836619ee-c877-45d9-b718-ab0eea4ed172"
response = requests.get(matchUrl)
matchlist = response.json()

try:
    totGames = matchlist['totalGames']
except:
    totGames = 0

if totGames > 0:
    for stuff in range(0, 2):
        if len(chGames) < 2:
            chGames.append(matchlist['matches'][stuff]['gameId'])
            

"""
matches = [[], [], [], [], [], [], []]

for i in range(0, len(sumIds)):
    for j in range(0, len(sumIds[i])):
        sumId = sumIds[i][j]
        matchTot = len(matches[i])
        if matchTot < caps[i]:
            time.sleep(2)
            matchUrl = "https://na1.api.riotgames.com/lol/match/v3/matchlists/by-account/"+str(sumId)+"?api_key=836619ee-c877-45d9-b718-ab0eea4ed172"
            response = requests.get(matchUrl)
            matchlist = response.json()
            
            try:
                totGames = matchlist['totalGames']
            except:
                totGames = 0
            
            if totGames > 0:
                for stuff in range(0, totGames):
                    if len(matches[i]) < caps[i]:
                        matches[i].append(matchlist['matches'][stuff]['gameId'])
"""