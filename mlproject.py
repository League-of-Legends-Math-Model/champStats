# -*- coding: utf-8 -*-
"""
Created on Sun Apr 30 22:44:51 2017

@author: Jon-Michael
"""

from openpyxl import load_workbook
from copy import copy, deepcopy
import numpy as np

workbook = load_workbook("seedStatsData.xlsx", data_only=True)
worksheet = workbook.worksheets[0]

row_count = worksheet.max_row


nte = True
row = 0

tops = []
jungles = []
mids = []
adcs = []
supps = []

for row in worksheet.iter_rows():
    holder = []
    for cell in row:
        holder.append(cell.internal_value)
    
    subHolder = []
    subHolder.append(float(holder[1]))  #RANKING
    subHolder.append(holder[12]) #DAMAGE DEALT PARTICIPATION
    subHolder.append(holder[13]) #KILL PARTICIPATION
    subHolder.append(holder[17]) #CS PER MINUTE
    subHolder.append(holder[18]) #DAMAGE PER MINUTE
    subHolder.append(holder[19]) #WARDS PER MINUTE
    subHolder.append(holder[21]) #DURABILITY
    subHolder.append(holder[22]) #DAMAGE EFFECTIVENESS
    subHolder.append(holder[23]) #KDA PER MINUTE
    
    if holder[10] == "0":
        tops.append(subHolder)
        
    if holder[10] == "1":
        jungles.append(subHolder)
        
    if holder[10] == "2":
        mids.append(subHolder)
    
    if holder[10] == "3":
        if holder[11] == "2" or holder[11] == "3":
            supps.append(subHolder)    
        else:
            adcs.append(subHolder)

"""
REGRESSION
"""
toppers = np.ndarray(shape = (len(tops),len(tops[0])))
for rd in range(0, len(tops)):
    for dr in range(0, len(tops[0])):
        toppers[rd][dr] = tops[rd][dr]

ytp = deepcopy(toppers)

for rd in range(0, len(ytp)):
    ytp[rd][0] = 1

ytptrans = deepcopy(ytp.transpose())

mult = ytptrans.dot(ytp)

inverse = np.linalg.inv(mult)

ytpplus = inverse.dot(ytptrans)

tpexpect = np.ndarray(shape = (len(tops), 1))
for rd in range(0, len(tops)):
    tpexpect[rd][0] = tops[rd][0]

topreg = ytpplus.dot(tpexpect)

