# -*- coding: utf-8 -*-
"""
Created on Sat May  6 08:17:05 2017

@author: Jon-Michael
"""

from openpyxl import load_workbook
from copy import copy, deepcopy
import numpy as np
import requests
import json
import time

matches = [[], [], [], [], [], [], []]
workbook = load_workbook("matches.xlsx", data_only=True)

for i in range(0, 7):
    worksheet = workbook.worksheets[i]
    rowCount = worksheet.max_row
    
    for row in worksheet.iter_rows():
        for cell in row:
            matches[i].append(cell.internal_value)
    
rawData = []

#toMatches = 1
matchFails = 0
for i in range(0, len(matches)):
    for j in range(0, len(matches[i])):
        holder = []
        time.sleep(2)
        matchId = matches[i][j]
        matchUrl = "https://na1.api.riotgames.com/lol/match/v3/matches/"+str(matchId)+"?api_key=836619ee-c877-45d9-b718-ab0eea4ed172"
        response = requests.get(matchUrl)
        matchData = response.json()
        try:   
            partIdent = matchData['participantIdentities']
            players = matchData['participants']
            gameLength = matchData['gameDuration']
            for k in range(0, 10):
                holder.append([partIdent[k]['participantId'], partIdent[k]['player']['accountId'], gameLength])
                teamData = [[0, 0, 0, 0], [0, 0, 0, 0]]
            for k in range(0, 10):
                timeline = players[k]['timeline']
                stats = players[k]['stats']
                pId = players[k]['participantId'] - 1
                tId = players[k]['teamId']
                holder[pId].append(tId)
                holder[pId].append(players[k]['highestAchievedSeasonTier'])
                holder[pId].append(timeline['lane'])
                holder[pId].append(timeline['role'])
                #KDA
                kills = stats['kills']
                if tId == 100:
                    teamData[0][0] += kills
                else:
                    teamData[1][0] += kills
                holder[pId].append(kills)
                holder[pId].append(stats['deaths'])
                holder[pId].append(stats['assists'])
                holder[pId].append(stats['turretKills'])
                #DAMAGE -> OUTPUT
                tDtC = stats['totalDamageDealtToChampions']
                tDD = stats['totalDamageDealt']
                if tId == 100:
                    teamData[0][1] += tDtC
                    teamData[0][2] += tDD
                else:
                    teamData[1][1] += tDtC
                    teamData[1][2] += tDD

                holder[pId].append(tDtC)
                holder[pId].append(tDD)
                holder[pId].append(stats['damageDealtToObjectives'])
                holder[pId].append(stats['damageDealtToTurrets'])
                #DAMAGE -> INPUT
                holder[pId].append(stats['damageSelfMitigated'])
                holder[pId].append(stats['totalHeal'])
                tDT = stats['totalDamageTaken']
                if tId == 100:
                    teamData[0][3] += tDT
                else:
                    teamData[1][3] += tDT
                holder[pId].append(tDT)
                holder[pId].append(stats['totalUnitsHealed'])
                #CC
                holder[pId].append(stats['totalTimeCrowdControlDealt'])
                holder[pId].append(stats['timeCCingOthers'])
                #MINIONS
                holder[pId].append(stats['totalMinionsKilled'])
                holder[pId].append(stats['neutralMinionsKilled'])
                holder[pId].append(stats['neutralMinionsKilledTeamJungle'])
                holder[pId].append(stats['neutralMinionsKilledEnemyJungle'])
                #VISION
                holder[pId].append(stats['visionScore'])
                holder[pId].append(stats['wardsPlaced'])
                holder[pId].append(stats['wardsKilled'])
                holder[pId].append(stats['visionWardsBoughtInGame'])
                holder[pId].append(stats['sightWardsBoughtInGame'])
                #GOLD
                holder[pId].append(stats['goldEarned'])
                holder[pId].append(stats['goldSpent'])
                #DURATION
                holder[pId].append(stats['longestTimeSpentLiving'])
                
            for k in range(0, 10):
                tId = players[k]['teamId']
                pId = players[k]['participantId'] - 1
                team = 1
                if tId == 100:
                    team = 0
                for l in range(0, 4):
                    holder[pId].append(teamData[team][l])
    
            for k in range(0, 10):
                rawData.append(holder[k])
        except:
            matchFails += 1
            
            